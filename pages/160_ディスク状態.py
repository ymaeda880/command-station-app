# pages/160_ディスク状態.py
from __future__ import annotations

# ============================================================
# imports
# ============================================================
import io
import platform
import re
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st

# ============================================================
# sys.path 調整（common_lib を import 可能に）
# ============================================================
_THIS = Path(__file__).resolve()
PROJECTS_ROOT = _THIS.parents[3]

if str(PROJECTS_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECTS_ROOT))

# ============================================================
# common_lib imports
# ============================================================
from common_lib.storage.external_ssd_root import resolve_storage_subdir_root  # noqa: E402

# ============================================================
# third-party imports
# ============================================================
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

# ============================================================
# app local imports
# ============================================================
from lib.cmd_utils import run_safe  # noqa: E402


# ============================================================
# Page
# ============================================================
st.set_page_config(page_title="💽 ディスク状態", page_icon="💽", layout="wide")
st.title("💽 ディスク状態")
st.caption("ディスク容量・マウント情報・Storages / InBoxStorages の容量を確認します。")


# ============================================================
# Paths
# ============================================================
STORAGE_ROOT = resolve_storage_subdir_root(PROJECTS_ROOT, subdir="Storages")
INBOX_ROOT = resolve_storage_subdir_root(PROJECTS_ROOT, subdir="InBoxStorages")


# ============================================================
# constants
# ============================================================
WARN_USE_PCT = 80.0
ERROR_USE_PCT = 90.0


# ============================================================
# size parse helpers
# ============================================================
_SIZE_RE = re.compile(
    r"^\s*(?P<num>\d+(?:\.\d+)?)\s*(?P<unit>[KMGTP]?i?B?)\s*$",
    re.IGNORECASE,
)

_UNIT_MAP = {
    "": 1,
    "B": 1,
    "K": 1000,
    "KB": 1000,
    "M": 1000**2,
    "MB": 1000**2,
    "G": 1000**3,
    "GB": 1000**3,
    "T": 1000**4,
    "TB": 1000**4,
    "P": 1000**5,
    "PB": 1000**5,
    "KI": 1024,
    "KIB": 1024,
    "MI": 1024**2,
    "MIB": 1024**2,
    "GI": 1024**3,
    "GIB": 1024**3,
    "TI": 1024**4,
    "TIB": 1024**4,
    "PI": 1024**5,
    "PIB": 1024**5,
}


# ============================================================
# size parse
# ============================================================
def _parse_size_to_bytes(s: str) -> Optional[int]:
    if s is None:
        return None

    s = str(s).strip()
    if not s:
        return None

    m = _SIZE_RE.match(s)
    if not m:
        return None

    num = float(m.group("num"))
    unit = (m.group("unit") or "").upper()

    if unit.endswith("I") and not unit.endswith("IB"):
        unit = unit + "B"

    mult = _UNIT_MAP.get(unit)
    if mult is None:
        return None

    return int(num * mult)


# ============================================================
# use percent parse
# ============================================================
def _parse_use_pct(s: str) -> Optional[int]:
    if s is None:
        return None

    m = re.match(r"^\s*(\d+)\s*%\s*$", str(s))
    return int(m.group(1)) if m else None


# ============================================================
# display helpers
# ============================================================
def _fmt_gib(n_bytes: int) -> str:
    return f"{n_bytes / (1024**3):,.2f} GiB"


def _fmt_pct(x: float) -> str:
    return f"{x:.1f}%"


def _truncate_middle(text: str, max_len: int = 40) -> str:
    text = "" if text is None else str(text)

    if len(text) <= max_len:
        return text

    head = max(10, (max_len - 1) // 2)
    tail = max(10, max_len - head - 1)

    return text[:head] + "…" + text[-tail:]


# ============================================================
# status judge
# ============================================================
def _judge_use_pct(use_pct: Optional[float]) -> Tuple[str, str]:
    if use_pct is None:
        return "不明", "info"

    if use_pct >= ERROR_USE_PCT:
        return "🔴 要対応", "error"

    if use_pct >= WARN_USE_PCT:
        return "⚠️ 注意", "warning"

    return "✅ 正常", "success"

# ============================================================
# df -h 判定対象外
# ============================================================
def _is_df_h_ignore_row(mounted_on: str, filesystem: str) -> bool:
    mounted_on = str(mounted_on or "")
    filesystem = str(filesystem or "")

    if mounted_on == "/dev":
        return True

    if filesystem in {"devfs", "map"}:
        return True

    if mounted_on.startswith("/System/Volumes/VM"):
        return True

    if mounted_on.startswith("/System/Volumes/Preboot"):
        return True

    if mounted_on.startswith("/System/Volumes/Update"):
        return True

    return False

# ============================================================
# overall judge
# ============================================================
def _overall_status(levels: List[str]) -> Tuple[str, str]:
    if "error" in levels:
        return "🔴 要対応", "error"

    if "warning" in levels:
        return "⚠️ 注意", "warning"

    return "✅ 正常", "success"


# ============================================================
# df -h parse
# ============================================================
def _parse_df_h_table(df_out: str) -> pd.DataFrame:
    lines = [ln.rstrip("\n") for ln in (df_out or "").splitlines() if ln.strip()]
    if not lines:
        return pd.DataFrame()

    header = lines[0].replace("Mounted on", "Mounted_on")
    cols = header.split()

    rows: List[Dict[str, str]] = []

    for ln in lines[1:]:
        parts = ln.split()

        if len(parts) < len(cols):
            continue

        if cols[-1] == "Mounted_on" and len(parts) > len(cols):
            parts = parts[: len(cols) - 1] + [" ".join(parts[len(cols) - 1 :])]

        if len(parts) > len(cols):
            parts = parts[: len(cols) - 1] + [" ".join(parts[len(cols) - 1 :])]

        row = {c: parts[i] if i < len(parts) else "" for i, c in enumerate(cols)}
        rows.append(row)

    df = pd.DataFrame(rows)

    if "Mounted_on" in df.columns:
        df = df.rename(columns={"Mounted_on": "Mounted on"})

    if "Capacity" in df.columns:
        df = df.rename(columns={"Capacity": "Use%"})

    return df


# ============================================================
# df -h view
# ============================================================
def _build_df_view(
    df_raw: pd.DataFrame,
    fs_len: int,
    mount_len: int,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if df_raw.empty:
        return df_raw, pd.DataFrame(), pd.DataFrame()

    df = df_raw.copy()

    fs_col = "Filesystem"
    mount_col = "Mounted on"
    use_col = "Use%" if "Use%" in df.columns else None

    df["_use_pct"] = df[use_col].apply(_parse_use_pct) if use_col else None

    df["_ignore_judge"] = df.apply(
        lambda r: _is_df_h_ignore_row(
            mounted_on=r.get(mount_col, ""),
            filesystem=r.get(fs_col, ""),
        ),
        axis=1,
    )

    df["_judge"], df["_level"] = zip(*df["_use_pct"].apply(_judge_use_pct))

    df.loc[df["_ignore_judge"], "_judge"] = "対象外"
    df.loc[df["_ignore_judge"], "_level"] = "ignore"

    df["_filesystem_full"] = df[fs_col].astype(str)
    df["_mounted_full"] = df[mount_col].astype(str)

    df["Filesystem"] = df["_filesystem_full"].apply(lambda s: _truncate_middle(s, fs_len))
    df["Mounted on"] = df["_mounted_full"].apply(lambda s: _truncate_middle(s, mount_len))

    show_cols = ["Mounted on", "Filesystem"]

    for c in ("Size", "Used", "Avail", "Use%"):
        if c in df.columns:
            show_cols.append(c)

    show_cols.append("判定")

    df["判定"] = df["_judge"]

    view = df[show_cols].copy()
    view["_is_root"] = (df["_mounted_full"] == "/").astype(int)
    view["_use_sort"] = df["_use_pct"].fillna(-1)

    view = (
        view.sort_values(by=["_is_root", "_use_sort"], ascending=[False, False])
        .drop(columns=["_is_root", "_use_sort"])
        .reset_index(drop=True)
    )

    mask = (df["_filesystem_full"].str.len() > fs_len) | (df["_mounted_full"].str.len() > mount_len)

    full_map = pd.DataFrame(
        {
            "Mounted on（表示）": df.loc[mask, "Mounted on"].values,
            "Mounted on（フル）": df.loc[mask, "_mounted_full"].values,
            "Filesystem（表示）": df.loc[mask, "Filesystem"].values,
            "Filesystem（フル）": df.loc[mask, "_filesystem_full"].values,
        }
    )

    issues = df[
        (~df["_ignore_judge"])
        & (df["_level"].isin(["warning", "error"]))
    ].copy()

    if not issues.empty:
        issue_cols = ["_judge", "_mounted_full", "_filesystem_full"]

        for c in ("Size", "Used", "Avail", "Use%"):
            if c in issues.columns:
                issue_cols.append(c)

        issues = issues[issue_cols].rename(
            columns={
                "_judge": "判定",
                "_mounted_full": "Mounted on",
                "_filesystem_full": "Filesystem",
            }
        )

    return view, full_map, issues


# ============================================================
# subprocess helper
# ============================================================
def _run_subprocess_safe(args: List[str], timeout_sec: int = 300) -> Tuple[int, str, str]:
    try:
        p = subprocess.run(
            args,
            capture_output=True,
            text=True,
            timeout=timeout_sec,
            check=False,
        )
        return p.returncode, p.stdout or "", p.stderr or ""

    except subprocess.TimeoutExpired:
        return 124, "", f"Timeout: {' '.join(args)}"

    except Exception as e:
        return 1, "", f"Failed: {' '.join(args)}\n{e}"


# ============================================================
# du depth1
# ============================================================
def _du_depth1_kb(parent: Path, timeout_sec: int = 900) -> pd.DataFrame:
    parent = parent.resolve()

    if platform.system() == "Darwin":
        args = ["du", "-k", "-d", "1", str(parent)]
    else:
        args = ["du", "-k", "--max-depth=1", str(parent)]

    code, out, err = _run_subprocess_safe(args, timeout_sec=timeout_sec)

    if code != 0:
        raise RuntimeError(f"du failed (code={code})\n{err}".strip())

    rows: List[Dict[str, object]] = []

    for ln in out.splitlines():
        ln = ln.strip()

        if not ln:
            continue

        parts = ln.split("\t", 1)

        if len(parts) != 2:
            parts2 = ln.split(None, 1)

            if len(parts2) != 2:
                continue

            parts = [parts2[0], parts2[1]]

        try:
            size_kb = int(parts[0])
        except Exception:
            continue

        p = Path(parts[1]).resolve()

        if p == parent:
            continue

        if p.parent != parent:
            continue

        if not p.exists() or not p.is_dir():
            continue

        rows.append(
            {
                "name": p.name,
                "path": str(p),
                "size_kb": size_kb,
                "size_bytes": size_kb * 1024,
            }
        )

    df = pd.DataFrame(rows)

    if df.empty:
        return df

    df = df.sort_values(by=["size_bytes"], ascending=False).reset_index(drop=True)

    return df


# ============================================================
# du total
# ============================================================
def _du_total_kb(path: Path, timeout_sec: int = 600) -> int:
    args = ["du", "-sk", str(path)]
    code, out, err = _run_subprocess_safe(args, timeout_sec=timeout_sec)

    if code != 0:
        raise RuntimeError(f"du -sk failed (code={code})\n{err}".strip())

    ln = out.strip().splitlines()[0]
    parts = ln.split("\t", 1)

    if len(parts) == 1:
        parts = ln.split(None, 1)

    return int(parts[0])


# ============================================================
# disk usage row
# ============================================================
def _disk_usage_row(label: str, p: Path) -> Dict[str, object]:
    total, used, free = shutil.disk_usage(p)
    use_pct = (used / total * 100.0) if total else 0.0
    judge, level = _judge_use_pct(use_pct)

    return {
        "label": label,
        "path": str(p),
        "fs_total_bytes": int(total),
        "fs_used_bytes": int(used),
        "fs_free_bytes": int(free),
        "fs_use_pct": float(use_pct),
        "judge": judge,
        "level": level,
    }


# ============================================================
# summary build
# ============================================================
def _build_summary(
    storage_root: Path,
    inbox_root: Path,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows = []

    for label, p in [("Storages", storage_root), ("InBoxStorages", inbox_root)]:
        total, used, free = shutil.disk_usage(p)
        use_pct = (used / total * 100.0) if total else 0.0
        dir_kb = _du_total_kb(p)
        judge, level = _judge_use_pct(use_pct)

        rows.append(
            {
                "label": label,
                "path": str(p),
                "fs_total_bytes": int(total),
                "fs_used_bytes": int(used),
                "fs_free_bytes": int(free),
                "fs_use_pct": float(use_pct),
                "dir_bytes": int(dir_kb * 1024),
                "judge": judge,
                "level": level,
                "measured_at": now,
            }
        )

    summary_raw = pd.DataFrame(rows)

    summary_raw["_fs_total_gib"] = summary_raw["fs_total_bytes"] / (1024**3)
    summary_raw["_fs_used_gib"] = summary_raw["fs_used_bytes"] / (1024**3)
    summary_raw["_fs_free_gib"] = summary_raw["fs_free_bytes"] / (1024**3)
    summary_raw["_dir_gib"] = summary_raw["dir_bytes"] / (1024**3)

    summary_view = pd.DataFrame(
        {
            "Target": summary_raw["label"],
            "判定": summary_raw["judge"],
            "Path": summary_raw["path"],
            "FS Total": summary_raw["fs_total_bytes"].map(_fmt_gib),
            "FS Used": summary_raw["fs_used_bytes"].map(_fmt_gib),
            "FS Free": summary_raw["fs_free_bytes"].map(_fmt_gib),
            "FS Use%": summary_raw["fs_use_pct"].map(_fmt_pct),
            "DIR Size": summary_raw["dir_bytes"].map(_fmt_gib),
            "Measured at": summary_raw["measured_at"],
        }
    )

    return summary_view, summary_raw


# ============================================================
# Excel bytes
# ============================================================
def _excel_bytes(
    summary_raw: pd.DataFrame,
    stor_l1: pd.DataFrame,
    inbox_l1: pd.DataFrame,
) -> bytes:
    wb = Workbook()

    # ------------------------------------------------------------
    # header style
    # ------------------------------------------------------------
    def _apply_header(ws, row: int, cols: int):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center")

    # ------------------------------------------------------------
    # auto size
    # ------------------------------------------------------------
    def _autosize(ws):
        for col in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col)

            for row in range(1, ws.max_row + 1):
                v = ws.cell(row=row, column=col).value

                if v is None:
                    continue

                max_len = max(max_len, len(str(v)))

            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)

    # ------------------------------------------------------------
    # Summary
    # ------------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"

    cols = [
        "label",
        "path",
        "fs_total_bytes",
        "fs_used_bytes",
        "fs_free_bytes",
        "fs_use_pct",
        "dir_bytes",
        "judge",
        "level",
        "measured_at",
        "_fs_total_gib",
        "_fs_used_gib",
        "_fs_free_gib",
        "_dir_gib",
    ]

    ws.append(cols)
    _apply_header(ws, 1, len(cols))

    for _, r in summary_raw.iterrows():
        ws.append([r.get(c) for c in cols])

    ws.freeze_panes = "A2"
    _autosize(ws)

    # ------------------------------------------------------------
    # Storages L1
    # ------------------------------------------------------------
    ws2 = wb.create_sheet("Storages_L1")

    if stor_l1 is None or stor_l1.empty:
        ws2.append(["(no rows)"])
    else:
        cols2 = ["name", "path", "size_gib", "size_bytes", "size_kb"]
        ws2.append(cols2)
        _apply_header(ws2, 1, len(cols2))

        for _, r in stor_l1.iterrows():
            ws2.append(
                [
                    r["name"],
                    r["path"],
                    round(r["size_bytes"] / (1024**3), 3),
                    int(r["size_bytes"]),
                    int(r["size_kb"]),
                ]
            )

        ws2.freeze_panes = "A2"
        _autosize(ws2)

    # ------------------------------------------------------------
    # InBoxStorages L1
    # ------------------------------------------------------------
    ws3 = wb.create_sheet("InBoxStorages_L1")

    if inbox_l1 is None or inbox_l1.empty:
        ws3.append(["(no rows)"])
    else:
        cols3 = ["name", "path", "size_gib", "size_bytes", "size_kb"]
        ws3.append(cols3)
        _apply_header(ws3, 1, len(cols3))

        for _, r in inbox_l1.iterrows():
            ws3.append(
                [
                    r["name"],
                    r["path"],
                    round(r["size_bytes"] / (1024**3), 3),
                    int(r["size_bytes"]),
                    int(r["size_kb"]),
                ]
            )

        ws3.freeze_panes = "A2"
        _autosize(ws3)

    # ------------------------------------------------------------
    # Meta
    # ------------------------------------------------------------
    ws4 = wb.create_sheet("Meta")
    ws4.append(["measured_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws4.append(["projects_root", str(PROJECTS_ROOT)])
    ws4.append(["storage_root", str(STORAGE_ROOT)])
    ws4.append(["inbox_root", str(INBOX_ROOT)])
    ws4.append(["platform", platform.platform()])

    bio = io.BytesIO()
    wb.save(bio)

    return bio.getvalue()


# ============================================================
# render status message
# ============================================================
def _render_status_box(status: str, level: str, detail: str) -> None:
    if level == "error":
        st.error(f"{status}  {detail}")
    elif level == "warning":
        st.warning(f"{status}  {detail}")
    elif level == "success":
        st.success(f"{status}  {detail}")
    else:
        st.info(f"{status}  {detail}")


# ============================================================
# root disk summary
# ============================================================
def _build_root_summary() -> Dict[str, object]:
    row = _disk_usage_row("Mac本体 /", Path("/"))

    return row


# ============================================================
# session state init
# ============================================================
if "stor_summary_view" not in st.session_state:
    st.session_state["stor_summary_view"] = None
    st.session_state["stor_summary_raw"] = None
    st.session_state["stor_l1"] = None
    st.session_state["inbox_l1"] = None
    st.session_state["stor_measured_at"] = None


# ============================================================
# Section: current status
# ============================================================
st.header("現在の状態")

root_summary = _build_root_summary()
root_judge = str(root_summary["judge"])
root_level = str(root_summary["level"])

summary_raw_current = st.session_state.get("stor_summary_raw")

levels_for_overall = [root_level]

if isinstance(summary_raw_current, pd.DataFrame) and not summary_raw_current.empty:
    levels_for_overall.extend(summary_raw_current["level"].astype(str).tolist())

overall_judge, overall_level = _overall_status(levels_for_overall)

_render_status_box(
    overall_judge,
    overall_level,
    "ディスク使用率を基準にした総合判定です。",
)

st.caption(
    f"判定基準：使用率 {WARN_USE_PCT:.0f}% 以上で注意、"
    f"{ERROR_USE_PCT:.0f}% 以上で要対応。"
)

# ============================================================
# Section: important cards
# ============================================================
card_cols = st.columns(3)

with card_cols[0]:
    st.subheader("Mac本体 /")
    st.metric(
        label="使用率",
        value=_fmt_pct(float(root_summary["fs_use_pct"])),
        delta=root_judge,
    )
    st.caption(f"総容量: {_fmt_gib(int(root_summary['fs_total_bytes']))}")
    st.caption(f"使用済: {_fmt_gib(int(root_summary['fs_used_bytes']))}")
    st.caption(f"空き: {_fmt_gib(int(root_summary['fs_free_bytes']))}")

with card_cols[1]:
    st.subheader("Storages")
    if isinstance(summary_raw_current, pd.DataFrame) and not summary_raw_current.empty:
        stor_row = summary_raw_current[summary_raw_current["label"] == "Storages"]
        if not stor_row.empty:
            r = stor_row.iloc[0]
            st.metric("フォルダー容量", _fmt_gib(int(r["dir_bytes"])), str(r["judge"]))
            st.caption("この値は Storages フォルダー自体が使用している容量です。")
            st.caption(f"保存先ディスク使用率: {_fmt_pct(float(r['fs_use_pct']))}")
            st.caption(f"保存先: {r['path']}")
        else:
            st.info("未計測です。")
    else:
        st.info("未計測です。")
        st.caption("下の「計測を実行」で表示します。")

with card_cols[2]:
    st.subheader("InBoxStorages")

    if isinstance(summary_raw_current, pd.DataFrame) and not summary_raw_current.empty:
        inbox_row = summary_raw_current[
            summary_raw_current["label"] == "InBoxStorages"
        ]

        if not inbox_row.empty:
            r = inbox_row.iloc[0]

            st.metric(
                "フォルダー容量",
                _fmt_gib(int(r["dir_bytes"])),
                str(r["judge"]),
            )

            st.caption(
                "この値は InBoxStorages フォルダー自体が使用している容量です。"
            )

            st.caption(
                "保存先ディスク使用率: "
                f"{_fmt_pct(float(r['fs_use_pct']))}"
            )

            st.caption(f"保存先: {r['path']}")


# ============================================================
# Section: df -h
# ============================================================
st.divider()
st.header("ディスク使用率一覧")

code, out, err = run_safe("df -h")

if err:
    st.error(err)

df_raw = pd.DataFrame()
view_df = pd.DataFrame()
full_df = pd.DataFrame()
issue_df = pd.DataFrame()

if out:
    df_raw = _parse_df_h_table(out)

    if df_raw.empty:
        st.warning("`df -h` の出力を表に変換できませんでした。")
    else:
        view_df, full_df, issue_df = _build_df_view(df_raw, fs_len=44, mount_len=44)

        if issue_df.empty:
            st.success("注意・要対応のディスクはありません。")
        else:
            st.warning("注意・要対応のディスクがあります。")
            st.dataframe(issue_df, hide_index=True)

        with st.expander("df -h 詳細表", expanded=False):
            c1, c2 = st.columns(2)

            with c1:
                fs_len = st.slider("Filesystem 表示長", 24, 80, 44, step=2)

            with c2:
                mount_len = st.slider("Mounted on 表示長", 24, 80, 44, step=2)

            view_df, full_df, issue_df = _build_df_view(df_raw, fs_len, mount_len)
            st.dataframe(view_df, hide_index=True)

            if not full_df.empty:
                with st.expander("省略されたフル文字列（コピー用）", expanded=False):
                    st.dataframe(full_df, hide_index=True)

        with st.expander("生ログ（df -h）", expanded=False):
            st.code(out, language="bash")
else:
    st.info("`df -h` の出力がありません。")


# ============================================================
# Section: diskutil list
# ============================================================
st.divider()

with st.expander("diskutil list 生ログ（macOS限定）", expanded=False):
    if platform.system() == "Darwin":
        code2, out2, err2 = run_safe("diskutil list")

        if err2:
            st.error(err2)

        if out2:
            st.code(out2, language="bash")
        else:
            st.info("`diskutil list` の出力がありません。")
    else:
        st.info("このコマンドは macOS 専用です。")


# ============================================================
# Section: Storages / InBoxStorages scan
# ============================================================
st.divider()
st.header("Storages / InBoxStorages 容量")

st.caption("FS容量はディスク全体の容量、DIR容量はそのフォルダ自体が使用している容量です。")
st.caption("直下フォルダ内訳は `du` で計測するため、容量が大きい場合は時間がかかります。")

# ============================================================
# scan settings
# ============================================================
cA, cB, cC = st.columns(3)

with cA:
    max_rows = st.number_input(
        "直下フォルダ表示 上位N件",
        min_value=10,
        max_value=2000,
        value=200,
        step=10,
    )

with cB:
    du_timeout = st.number_input(
        "du タイムアウト（秒）",
        min_value=30,
        max_value=3600,
        value=900,
        step=30,
    )

with cC:
    show_hidden = st.checkbox(
        "隠しフォルダ（.で始まる）も含める",
        value=False,
    )


# ============================================================
# scan buttons
# ============================================================
btn_cols = st.columns(2)

with btn_cols[0]:
    run_scan = st.button("計測を実行", key="run_storage_scan")

with btn_cols[1]:
    clear_scan = st.button("結果をクリア", key="clear_storage_scan")


# ============================================================
# clear scan
# ============================================================
if clear_scan:
    st.session_state["stor_summary_view"] = None
    st.session_state["stor_summary_raw"] = None
    st.session_state["stor_l1"] = None
    st.session_state["inbox_l1"] = None
    st.session_state["stor_measured_at"] = None
    st.rerun()


# ============================================================
# run scan
# ============================================================
if run_scan:
    with st.spinner("サマリー（FS容量 + DIR容量）を計測中…"):
        summary_view, summary_raw = _build_summary(STORAGE_ROOT, INBOX_ROOT)

        st.session_state["stor_summary_view"] = summary_view
        st.session_state["stor_summary_raw"] = summary_raw
        st.session_state["stor_measured_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with st.spinner("Storages 直下フォルダを計測中…"):
        stor_l1 = _du_depth1_kb(STORAGE_ROOT, timeout_sec=int(du_timeout))

        if not show_hidden and not stor_l1.empty:
            stor_l1 = stor_l1[
                ~stor_l1["name"].astype(str).str.startswith(".")
            ].reset_index(drop=True)

        st.session_state["stor_l1"] = stor_l1

    with st.spinner("InBoxStorages 直下フォルダを計測中…"):
        inbox_l1 = _du_depth1_kb(INBOX_ROOT, timeout_sec=int(du_timeout))

        if not show_hidden and not inbox_l1.empty:
            inbox_l1 = inbox_l1[
                ~inbox_l1["name"].astype(str).str.startswith(".")
            ].reset_index(drop=True)

        st.session_state["inbox_l1"] = inbox_l1

    st.rerun()


# ============================================================
# scan result vars
# ============================================================
summary_view = st.session_state.get("stor_summary_view")
summary_raw = st.session_state.get("stor_summary_raw")
stor_l1 = st.session_state.get("stor_l1")
inbox_l1 = st.session_state.get("inbox_l1")
stor_measured_at = st.session_state.get("stor_measured_at")


# ============================================================
# render scan results
# ============================================================
if summary_view is not None:
    st.subheader("サマリー")

    if stor_measured_at:
        st.caption(f"最終計測: {stor_measured_at}")

    st.dataframe(summary_view, hide_index=True)

    # ------------------------------------------------------------
    # issues from Storages / InBoxStorages
    # ------------------------------------------------------------
    if isinstance(summary_raw, pd.DataFrame) and not summary_raw.empty:
        issue_summary = summary_raw[summary_raw["level"].isin(["warning", "error"])].copy()

        if issue_summary.empty:
            st.success("Storages / InBoxStorages のFS使用率に注意・要対応はありません。")
        else:
            st.warning("Storages / InBoxStorages に注意・要対応があります。")

            issue_view = pd.DataFrame(
                {
                    "判定": issue_summary["judge"],
                    "Target": issue_summary["label"],
                    "Path": issue_summary["path"],
                    "FS Use%": issue_summary["fs_use_pct"].map(_fmt_pct),
                    "FS Free": issue_summary["fs_free_bytes"].map(_fmt_gib),
                    "DIR Size": issue_summary["dir_bytes"].map(_fmt_gib),
                }
            )

            st.dataframe(issue_view, hide_index=True)

    # ------------------------------------------------------------
    # l1 tabs
    # ------------------------------------------------------------
    t1, t2 = st.tabs(["Storages（直下フォルダ）", "InBoxStorages（直下フォルダ）"])

    with t1:
        if stor_l1 is None or stor_l1.empty:
            st.info("直下フォルダが見つからないか、計測結果が空です。")
        else:
            view = stor_l1.copy()
            view["size_gib"] = view["size_bytes"].map(_fmt_gib)
            view = view[["name", "path", "size_gib", "size_bytes"]].head(int(max_rows))
            st.dataframe(view, hide_index=True)

    with t2:
        if inbox_l1 is None or inbox_l1.empty:
            st.info("直下フォルダが見つからないか、計測結果が空です。")
        else:
            view = inbox_l1.copy()
            view["size_gib"] = view["size_bytes"].map(_fmt_gib)
            view = view[["name", "path", "size_gib", "size_bytes"]].head(int(max_rows))
            st.dataframe(view, hide_index=True)

    # ------------------------------------------------------------
    # Excel download
    # ------------------------------------------------------------
    st.subheader("Excel ダウンロード")

    try:
        xlsx = _excel_bytes(summary_raw, stor_l1, inbox_l1)
        fname = f"storages_usage_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        st.download_button(
            label="Excel（.xlsx）をダウンロード",
            data=xlsx,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        st.error(f"Excel生成に失敗しました: {e}")

else:
    st.info("「計測を実行」を押すと、Storages / InBoxStorages の容量と直下フォルダ内訳を表示します。")


# ============================================================
# Section: arbitrary path disk usage
# ============================================================
st.divider()
st.header("任意パスの空き容量チェック")

target = st.text_input("対象パス", "/")

try:
    total, used, free = shutil.disk_usage(target)

    pct_used = (used / total * 100.0) if total else 0.0
    pct_free = 100.0 - pct_used

    judge, level = _judge_use_pct(pct_used)

    _render_status_box(
        judge,
        level,
        f"`{target}` の使用率は {_fmt_pct(pct_used)} です。",
    )

    col_a, col_b, col_c = st.columns(3)

    with col_a:
        st.metric("総容量", _fmt_gib(total))

    with col_b:
        st.metric("使用済", _fmt_gib(used), _fmt_pct(pct_used))

    with col_c:
        st.metric("空き", _fmt_gib(free), _fmt_pct(pct_free))

except Exception as e:
    st.error(f"取得に失敗しました: {e}")